import pandas as pd
import numpy as np
from openpyxl.styles import Font, PatternFill, Alignment, numbers

"""
@description 廣告素材分析腳本
@author AI Assistant
@date 2024
@version 1.1.1

功能說明：
- 分析優化目標為安裝及保留的廣告活動數據
- 按聯播網版位和素材類型分析
- 計算每季花費最多的Top3素材及相關指標
- 添加安裝相關指標
- 優化Excel格式和超連結
- 修正數值格式問題
"""

def calculate_metrics(group_data):
    """計算指標"""
    total_cost = group_data['費用'].sum()
    total_impressions = group_data['曝光'].sum()
    total_installs = group_data['安裝'].sum()
    
    metrics = {
        '花費總和': total_cost,
        '花費佔比': group_data['費用'].sum() / total_cost * 100 if total_cost > 0 else 0,
        '曝光總和': group_data['曝光'].sum(),
        '曝光佔比': group_data['曝光'].sum() / total_impressions * 100 if total_impressions > 0 else 0,
        'IR': group_data['安裝'].sum() / group_data['曝光'].sum() * 100 if group_data['曝光'].sum() > 0 else 0,
        'CPI': group_data['費用'].sum() / group_data['安裝'].sum() if group_data['安裝'].sum() > 0 else 0
    }
    return pd.Series(metrics)

def analyze_by_network_and_type(df, network, material_type):
    """按聯播網和素材類型分析"""
    # 篩選數據
    mask = (df['聯播網 (及搜尋夥伴)'] == network) & \
           (df['素材類型'] == material_type) & \
           (df['優化目標'].isin(['安裝', '保留']))
    
    data = df[mask].copy()
    
    # 添加季度列
    data['季度'] = data['月'].dt.to_period('Q')
    
    # 按季度分組
    quarterly_groups = data.groupby(['季度', '素材 ID', '素材名稱', 'url']).\
        agg({
            '費用': 'sum',
            '曝光': 'sum',
            '安裝': 'sum'
        }).reset_index()
    
    # 按季度排序
    quarterly_groups['季度'] = pd.PeriodIndex(quarterly_groups['季度'])
    quarterly_groups = quarterly_groups.sort_values('季度')
    
    top3_results = []
    worst3_results = []
    poor_performance_results = []
    
    # 追蹤已經出現過的素材
    seen_materials = set()
    
    # 對每個季度找出Top3和Worst3素材，以及表現不佳的新素材
    for quarter in quarterly_groups['季度'].unique():
        quarter_data = quarterly_groups[quarterly_groups['季度'] == quarter].copy()
        
        # 計算該季度的平均指標
        total_impressions = quarter_data['曝光'].sum()
        total_installs = quarter_data['安裝'].sum()
        total_cost = quarter_data['費用'].sum()
        
        # 計算平均IR和CPI
        avg_ir = total_installs / total_impressions if total_impressions > 0 else 0
        avg_cpi = total_cost / total_installs if total_installs > 0 else 0
        
        # 計算每個素材的IR
        quarter_data.loc[:, 'IR'] = quarter_data.apply(
            lambda x: (x['安裝'] / x['曝光']) if x['曝光'] > 0 else 0, axis=1
        )
        
        # 按花費排序取Top3
        top3 = quarter_data.nlargest(3, '費用')
        # 按IR排序取Worst3（排除IR為0的素材）
        worst3 = quarter_data[quarter_data['IR'] > 0].nsmallest(3, 'IR')
        
        # 找出該季度新上架且表現不佳的素材
        for _, row in quarter_data.iterrows():
            material_id = row['素材 ID']
            # 檢查是否為新素材（本季之前都沒有曝光過）
            if material_id not in seen_materials:
                cost_share = row['費用'] / total_cost if total_cost > 0 else 0
                ir = row['IR']
                # 檢查是否符合表現不佳的標準
                if cost_share < 0.01 or (avg_ir > 0 and ir < avg_ir * 0.5):
                    result = {
                        '季度': quarter,
                        '素材 ID': material_id,
                        '素材名稱': row['素材名稱'],
                        '花費總和': row['費用'],
                        '花費佔比': cost_share,
                        '曝光總和': row['曝光'],
                        '曝光佔比': (row['曝光'] / total_impressions) if total_impressions > 0 else 0,
                        '安裝總和': row['安裝'],
                        '安裝佔比': (row['安裝'] / total_installs) if total_installs > 0 else 0,
                        'IR': ir,
                        'CPI': row['費用'] / row['安裝'] if row['安裝'] > 0 else 0,
                        '平均IR': avg_ir,
                        '平均CPI': avg_cpi,
                        '問題類型': '花費佔比<1%' if cost_share < 0.01 else 'IR<平均值50%' if ir < avg_ir * 0.5 else '兩者皆是',
                        'url': row['url']
                    }
                    poor_performance_results.append(result)
            
            # 將當前素材加入已見過的集合
            seen_materials.add(material_id)
        
        # 為每個Top3素材計算指標
        for _, row in top3.iterrows():
            result = {
                '季度': quarter,
                '素材 ID': row['素材 ID'],
                '素材名稱': row['素材名稱'],
                '花費總和': row['費用'],
                '花費佔比': (row['費用'] / total_cost) if total_cost > 0 else 0,
                '曝光總和': row['曝光'],
                '曝光佔比': (row['曝光'] / total_impressions) if total_impressions > 0 else 0,
                '安裝總和': row['安裝'],
                '安裝佔比': (row['安裝'] / total_installs) if total_installs > 0 else 0,
                'IR': (row['安裝'] / row['曝光']) if row['曝光'] > 0 else 0,
                'CPI': row['費用'] / row['安裝'] if row['安裝'] > 0 else 0,
                '平均IR': avg_ir,
                '平均CPI': avg_cpi,
                'url': row['url']
            }
            top3_results.append(result)
            
        # 為每個Worst3素材計算指標
        for _, row in worst3.iterrows():
            result = {
                '季度': quarter,
                '素材 ID': row['素材 ID'],
                '素材名稱': row['素材名稱'],
                '花費總和': row['費用'],
                '花費佔比': (row['費用'] / total_cost) if total_cost > 0 else 0,
                '曝光總和': row['曝光'],
                '曝光佔比': (row['曝光'] / total_impressions) if total_impressions > 0 else 0,
                '安裝總和': row['安裝'],
                '安裝佔比': (row['安裝'] / total_installs) if total_installs > 0 else 0,
                'IR': (row['安裝'] / row['曝光']) if row['曝光'] > 0 else 0,
                'CPI': row['費用'] / row['安裝'] if row['安裝'] > 0 else 0,
                '平均IR': avg_ir,
                '平均CPI': avg_cpi,
                'url': row['url']
            }
            worst3_results.append(result)
    
    return pd.DataFrame(top3_results), pd.DataFrame(worst3_results), pd.DataFrame(poor_performance_results)

# 讀取數據
df = pd.read_csv('GG.csv', encoding='utf-8-sig')

# 將日期列轉換為日期格式
df['月'] = pd.to_datetime(df['月'])

# 分析各個版位和素材類型
display_youtube_top3, display_youtube_worst3, display_youtube_poor = analyze_by_network_and_type(df, 'Google 多媒體廣告聯播網', 'YouTube 影片')
display_image_top3, display_image_worst3, display_image_poor = analyze_by_network_and_type(df, 'Google 多媒體廣告聯播網', '橫向圖片')
youtube_youtube_top3, youtube_youtube_worst3, youtube_youtube_poor = analyze_by_network_and_type(df, 'YouTube', 'YouTube 影片')
youtube_image_top3, youtube_image_worst3, youtube_image_poor = analyze_by_network_and_type(df, 'YouTube', '橫向圖片')

# 保存到Excel並設置格式
with pd.ExcelWriter('material_analysis_result.xlsx', engine='openpyxl') as writer:
    # 保存數據
    # 聯播網_YouTube影片
    display_youtube_top3.to_excel(writer, sheet_name='聯播網_YouTube影片', index=False, startrow=0)
    display_youtube_worst3.to_excel(writer, sheet_name='聯播網_YouTube影片', index=False, startrow=len(display_youtube_top3) + 2)
    display_youtube_poor.to_excel(writer, sheet_name='聯播網_YouTube影片', index=False, startrow=len(display_youtube_top3) + len(display_youtube_worst3) + 4)
    
    # 聯播網_橫向圖片
    display_image_top3.to_excel(writer, sheet_name='聯播網_橫向圖片', index=False, startrow=0)
    display_image_worst3.to_excel(writer, sheet_name='聯播網_橫向圖片', index=False, startrow=len(display_image_top3) + 2)
    display_image_poor.to_excel(writer, sheet_name='聯播網_橫向圖片', index=False, startrow=len(display_image_top3) + len(display_image_worst3) + 4)
    
    # YouTube_YouTube影片
    youtube_youtube_top3.to_excel(writer, sheet_name='YouTube_YouTube影片', index=False, startrow=0)
    youtube_youtube_worst3.to_excel(writer, sheet_name='YouTube_YouTube影片', index=False, startrow=len(youtube_youtube_top3) + 2)
    youtube_youtube_poor.to_excel(writer, sheet_name='YouTube_YouTube影片', index=False, startrow=len(youtube_youtube_top3) + len(youtube_youtube_worst3) + 4)
    
    # YouTube_橫向圖片
    youtube_image_top3.to_excel(writer, sheet_name='YouTube_橫向圖片', index=False, startrow=0)
    youtube_image_worst3.to_excel(writer, sheet_name='YouTube_橫向圖片', index=False, startrow=len(youtube_image_top3) + 2)
    youtube_image_poor.to_excel(writer, sheet_name='YouTube_橫向圖片', index=False, startrow=len(youtube_image_top3) + len(youtube_image_worst3) + 4)
    
    # 設置每個工作表的格式
    workbook = writer.book
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # 獲取各個表格的行數
        if sheet_name == '聯播網_YouTube影片':
            top3_rows = len(display_youtube_top3)
            worst3_rows = len(display_youtube_worst3)
            poor_rows = len(display_youtube_poor)
        elif sheet_name == '聯播網_橫向圖片':
            top3_rows = len(display_image_top3)
            worst3_rows = len(display_image_worst3)
            poor_rows = len(display_image_poor)
        elif sheet_name == 'YouTube_YouTube影片':
            top3_rows = len(youtube_youtube_top3)
            worst3_rows = len(youtube_youtube_worst3)
            poor_rows = len(youtube_youtube_poor)
        else:
            top3_rows = len(youtube_image_top3)
            worst3_rows = len(youtube_image_worst3)
            poor_rows = len(youtube_image_poor)
        
        # 設置三個表格的標題行格式
        for row_num in [1, top3_rows + 3, top3_rows + worst3_rows + 5]:
            for cell in worksheet[row_num]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
        
        # 添加表格標題
        worksheet.cell(row=1, column=1, value='Top 3 (花費)').font = Font(bold=True, color='008000')
        worksheet.cell(row=top3_rows + 3, column=1, value='Worst 3 (IR)').font = Font(bold=True, color='FF0000')
        worksheet.cell(row=top3_rows + worst3_rows + 5, column=1, value='新素材問題清單').font = Font(bold=True, color='0000FF')
        
        # 設置數據格式
        for row in worksheet.iter_rows(min_row=2, max_row=top3_rows + worst3_rows + poor_rows + 5):
            for cell in row:
                if not cell.value:  # 跳過空行
                    continue
                column_name = worksheet.cell(row=1, column=cell.column).value
                if column_name in ['花費總和', '曝光總和', '安裝總和', 'CPI', '平均CPI']:
                    cell.number_format = '#,##0'
                elif column_name in ['花費佔比', '曝光佔比', '安裝佔比', 'IR', '平均IR']:
                    cell.number_format = '0.00%'
                elif column_name == 'url':
                    if cell.value and 'youtube.com' in str(cell.value):
                        cell.hyperlink = cell.value
                        cell.font = Font(color='0000FF', underline='single')
                elif column_name == '問題類型':
                    cell.font = Font(color='FF0000')
        
        # 調整列寬
        for column in worksheet.columns:
            max_length = 0
            column = list(column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

print("分析完成，結果已保存到 material_analysis_result.xlsx")

# 顯示第一個分析結果預覽
print("\nGoogle 多媒體廣告聯播網 - YouTube影片分析預覽：")
print("\nTop 3 (花費)：")
print(display_youtube_top3.to_string())
print("\nWorst 3 (IR)：")
print(display_youtube_worst3.to_string())
print("\n新素材問題清單：")
print(display_youtube_poor.to_string()) 